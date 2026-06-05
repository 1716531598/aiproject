from datetime import datetime, timedelta
from io import BytesIO

import openpyxl
from flask import Flask
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from models import IssueBug, IssueOverdueAnalysis, IssueProduct
from utils.db import Base


def _json_body(response):
    if isinstance(response, tuple):
        return response[0].get_json()
    return response.get_json()


def _setup_session(monkeypatch):
    from blueprints import issue_statistic

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    TestingSession = sessionmaker(bind=engine)
    monkeypatch.setattr(issue_statistic, "SessionLocal", TestingSession)
    return issue_statistic, TestingSession


def _seed_overdue_data(TestingSession):
    db = TestingSession()
    product = IssueProduct(name="产品A")
    db.add(product)
    db.commit()
    now = datetime.now()
    db.add_all(
        [
            IssueBug(
                bug_id="BUG-OD-1",
                product_id=product.id,
                title="超期40天",
                status="激活",
                affect_version="V1.0",
                created_date=now - timedelta(days=40),
            ),
            IssueBug(
                bug_id="BUG-OD-2",
                product_id=product.id,
                title="超期200天",
                status="激活",
                affect_version="V1.0",
                created_date=now - timedelta(days=200),
            ),
            IssueBug(
                bug_id="BUG-OD-3",
                product_id=product.id,
                title="超期400天",
                status="激活",
                affect_version="V1.0",
                created_date=now - timedelta(days=400),
            ),
            IssueBug(
                bug_id="BUG-OD-4",
                product_id=product.id,
                title="已解决不统计",
                status="已解决",
                affect_version="V1.0",
                created_date=now - timedelta(days=400),
            ),
        ]
    )
    db.commit()
    product_id = product.id
    db.close()
    return product_id


def test_overdue_query_analysis_and_export(monkeypatch):
    issue_statistic, TestingSession = _setup_session(monkeypatch)
    product_id = _seed_overdue_data(TestingSession)
    app = Flask(__name__)

    with app.test_request_context(json={}):
        body = _json_body(issue_statistic.overdue.__wrapped__())
    assert body["code"] == 200
    assert len(body["data"]) == 1
    row = body["data"][0]
    assert row["product_name"] == "产品A"
    assert row["version"] == "V1.0"
    assert row["overdue_total"] == 3
    assert row["overdue_180"] == 2
    assert row["overdue_360"] == 1
    assert [bug["bug_id"] for bug in row["bugs"]] == ["BUG-OD-3", "BUG-OD-2", "BUG-OD-1"]

    with app.test_request_context(
        json={
            "product_id": product_id,
            "version": "V1.0",
            "analysis": "历史版本遗留问题集中",
            "improvement": "建立版本清理计划",
        }
    ):
        body = _json_body(issue_statistic.overdue_analysis.__wrapped__())
    assert body["code"] == 200
    assert body["data"]["analysis"] == "历史版本遗留问题集中"

    db = TestingSession()
    assert db.query(IssueOverdueAnalysis).count() == 1
    db.close()

    response = issue_statistic.overdue_export.__wrapped__()
    workbook = openpyxl.load_workbook(BytesIO(response.get_data()))
    sheet = workbook.active
    assert sheet.cell(row=1, column=1).value == "产品"
    assert sheet.cell(row=2, column=1).value == "产品A"
    assert sheet.cell(row=2, column=3).value == 3
    assert sheet.cell(row=2, column=6).value == "历史版本遗留问题集中"


def test_issue_overdue_routes_are_registered(monkeypatch):
    import app as app_module

    monkeypatch.setattr(app_module, "init_db", lambda config: None)
    monkeypatch.setattr(app_module, "init_redis", lambda config: None)

    flask_app = app_module.create_app()
    rules = {rule.rule for rule in flask_app.url_map.iter_rules()}
    assert "/api/issue/statistics/overdue" in rules
    assert "/api/issue/statistics/overdue/analysis" in rules
    assert "/api/issue/statistics/overdue/export" in rules
