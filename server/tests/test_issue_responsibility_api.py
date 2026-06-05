from datetime import datetime
from io import BytesIO

import openpyxl
from flask import Flask
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from models import IssueAssessment, IssueBug, IssueProduct, IssueResponsibility, IssueStaff, IssueVersion, User
from utils.db import Base


def _json_body(response):
    if isinstance(response, tuple):
        return response[0].get_json()
    return response.get_json()


def _setup_session(monkeypatch):
    from blueprints import issue_responsibility

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    TestingSession = sessionmaker(bind=engine)
    monkeypatch.setattr(issue_responsibility, "SessionLocal", TestingSession)
    monkeypatch.setattr(issue_responsibility, "get_current_user", lambda: {"id": 1, "name": "zhangsan", "role_id": 1})
    return issue_responsibility, TestingSession


def _seed_data(TestingSession):
    db = TestingSession()
    user = User(id=1, name="zhangsan", password_hash="hash", role_id=1)
    product = IssueProduct(name="RayScan", description="")
    db.add_all([user, product])
    db.commit()
    version = IssueVersion(product_id=product.id, version="V1.0", status="开发中")
    staff = IssueStaff(name="张三", email="zhangsan@example.com", user_id=user.id, department="研发")
    staff2 = IssueStaff(name="李四", email="lisi@example.com", department="测试")
    db.add_all([version, staff, staff2])
    db.commit()
    db.add(IssueAssessment(product_id=product.id, version_id=version.id))
    db.add_all(
        [
            IssueBug(
                bug_id="BUG-1",
                product_id=product.id,
                title="登录失败",
                severity=1,
                affect_version="V1.0",
                created_date=datetime(2026, 6, 1, 10, 0, 0),
            ),
            IssueBug(
                bug_id="BUG-2",
                product_id=product.id,
                title="非考核版本",
                severity=2,
                affect_version="V2.0",
                created_date=datetime(2026, 6, 2, 10, 0, 0),
            ),
        ]
    )
    db.commit()
    ids = {
        "product_id": product.id,
        "bug_id": db.query(IssueBug).filter(IssueBug.bug_id == "BUG-1").first().id,
        "staff_id": staff.id,
        "staff2_id": staff2.id,
    }
    db.close()
    return ids


def test_responsibility_list_save_score_me_and_export(monkeypatch):
    issue_responsibility, TestingSession = _setup_session(monkeypatch)
    ids = _seed_data(TestingSession)
    app = Flask(__name__)

    with app.test_request_context(json={"page": 1, "pageSize": 20, "assigned": False}):
        body = _json_body(issue_responsibility.resp_list.__wrapped__())
    assert body["code"] == 200
    assert body["data"]["total"] == 1
    assert body["data"]["aaData"][0]["bug_id"] == "BUG-1"
    assert body["data"]["aaData"][0]["has_responsibility"] is False

    bad_payload = {
        "bug_id": ids["bug_id"],
        "records": [{"staff_id": ids["staff_id"], "role": "开发", "ratio": 0.5, "description": "根因"}],
    }
    with app.test_request_context(json=bad_payload):
        body = _json_body(issue_responsibility.resp_save.__wrapped__())
    assert body["code"] == 400
    assert body["msg"] == "责任占比合计不为 100%"

    payload = {
        "bug_id": ids["bug_id"],
        "year": 2026,
        "records": [
            {"staff_id": ids["staff_id"], "role": "开发", "ratio": 0.7, "description": "根因"},
            {"staff_id": ids["staff2_id"], "role": "测试", "ratio": 0.3, "description": "漏测"},
        ],
    }
    with app.test_request_context(json=payload):
        body = _json_body(issue_responsibility.resp_save.__wrapped__())
    assert body["code"] == 200

    with app.test_request_context(json={"page": 1, "pageSize": 20, "assigned": True}):
        body = _json_body(issue_responsibility.resp_list.__wrapped__())
    assert body["data"]["total"] == 1
    assert len(body["data"]["aaData"][0]["responsibilities"]) == 2

    with app.test_request_context(json={"year": 2026}):
        body = _json_body(issue_responsibility.score_query.__wrapped__())
    assert body["code"] == 200
    assert body["data"][0]["staff_name"] == "张三"
    assert body["data"][0]["raw_score"] == 0.07

    with app.test_request_context(json={"year": 2026}):
        body = _json_body(issue_responsibility.score_me.__wrapped__())
    assert body["code"] == 200
    assert body["data"]["total_score"] == 0.07
    assert body["data"]["records"][0]["bug_id"] == "BUG-1"

    with app.test_request_context(json={"year": 2026}):
        response = issue_responsibility.resp_export.__wrapped__()
    assert response.status_code == 200
    workbook = openpyxl.load_workbook(BytesIO(response.get_data()))
    sheet = workbook.active
    assert sheet.cell(row=1, column=1).value == "Bug 编号"
    assert sheet.cell(row=2, column=1).value == "BUG-1"

    db = TestingSession()
    assert db.query(IssueResponsibility).count() == 2
    db.close()


def test_issue_responsibility_routes_are_registered(monkeypatch):
    import app as app_module

    monkeypatch.setattr(app_module, "init_db", lambda config: None)
    monkeypatch.setattr(app_module, "init_redis", lambda config: None)

    flask_app = app_module.create_app()
    rules = {rule.rule for rule in flask_app.url_map.iter_rules()}

    assert "/api/issue/responsibilities/list" in rules
    assert "/api/issue/responsibilities/save" in rules
    assert "/api/issue/responsibilities/score/query" in rules
    assert "/api/issue/responsibilities/score/me" in rules
    assert "/api/issue/responsibilities/export" in rules
