from datetime import datetime
from io import BytesIO

import openpyxl
from flask import Flask
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from models import IssueBug, IssueProduct, IssueStaff
from utils.db import Base
from utils.issue_export import export_bugs_to_excel


def _json_body(response):
    if isinstance(response, tuple):
        return response[0].get_json()
    return response.get_json()


def _setup_session(monkeypatch):
    from blueprints import issue_bug

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    TestingSession = sessionmaker(bind=engine)
    monkeypatch.setattr(issue_bug, "SessionLocal", TestingSession)
    return issue_bug, TestingSession


def test_export_bugs_to_excel_writes_selected_columns():
    excel_bytes = export_bugs_to_excel(
        [{"bug_id": "BUG-1", "title": "登录失败", "product_name": "RayScan", "severity": 1}],
        columns=["bug_id", "title", "severity"],
    )

    workbook = openpyxl.load_workbook(BytesIO(excel_bytes))
    sheet = workbook.active
    assert sheet.title == "网上问题"
    assert [sheet.cell(row=1, column=i).value for i in range(1, 4)] == ["Bug 编号", "标题", "严重程度"]
    assert [sheet.cell(row=2, column=i).value for i in range(1, 4)] == ["BUG-1", "登录失败", 1]


def test_batch_assign_updates_selected_bugs(monkeypatch):
    issue_bug, TestingSession = _setup_session(monkeypatch)
    db = TestingSession()
    product = IssueProduct(name="RayScan", description="")
    staff = IssueStaff(name="张三", email="zhangsan@example.com")
    db.add_all([product, staff])
    db.commit()
    db.add_all(
        [
            IssueBug(bug_id="BUG-1", product_id=product.id, title="登录失败"),
            IssueBug(bug_id="BUG-2", product_id=product.id, title="样式异常"),
        ]
    )
    db.commit()
    bug_ids = [bug.id for bug in db.query(IssueBug).order_by(IssueBug.id).all()]
    staff_id = staff.id
    db.close()

    app = Flask(__name__)
    with app.test_request_context(json={"ids": bug_ids, "staff_id": staff_id}):
        body = _json_body(issue_bug.batch_assign.__wrapped__())

    assert body["code"] == 200
    assert "成功指派 2 个问题" in body["msg"]

    db = TestingSession()
    assigned = db.query(IssueBug).order_by(IssueBug.id).all()
    assert [item.staff_id for item in assigned] == [staff_id, staff_id]
    assert all(item.assign_time is not None for item in assigned)
    db.close()


def test_export_bugs_uses_query_filters_and_returns_xlsx(monkeypatch):
    issue_bug, TestingSession = _setup_session(monkeypatch)
    db = TestingSession()
    product = IssueProduct(name="RayScan", description="")
    db.add(product)
    db.commit()
    db.add_all(
        [
            IssueBug(
                bug_id="BUG-1",
                product_id=product.id,
                title="登录失败",
                severity=1,
                status="激活",
                created_date=datetime(2026, 6, 1, 10, 0, 0),
            ),
            IssueBug(
                bug_id="BUG-2",
                product_id=product.id,
                title="样式异常",
                severity=3,
                status="已解决",
                created_date=datetime(2026, 6, 2, 10, 0, 0),
            ),
        ]
    )
    db.commit()
    db.close()

    app = Flask(__name__)
    with app.test_request_context(json={"keyword": "登录", "columns": ["bug_id", "title", "product_name"]}):
        response = issue_bug.export_bugs.__wrapped__()

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = openpyxl.load_workbook(BytesIO(response.get_data()))
    sheet = workbook.active
    assert sheet.cell(row=2, column=1).value == "BUG-1"
    assert sheet.cell(row=3, column=1).value is None


def test_issue_bug_batch_export_routes_are_registered(monkeypatch):
    import app as app_module

    monkeypatch.setattr(app_module, "init_db", lambda config: None)
    monkeypatch.setattr(app_module, "init_redis", lambda config: None)

    flask_app = app_module.create_app()
    rules = {rule.rule for rule in flask_app.url_map.iter_rules()}

    assert "/api/issue/bugs/batch-assign" in rules
    assert "/api/issue/bugs/export" in rules
