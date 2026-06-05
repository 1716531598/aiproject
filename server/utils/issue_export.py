from io import BytesIO

import openpyxl
from openpyxl.styles import Font

BUG_EXPORT_COLUMNS = [
    ("bug_id", "Bug 编号"),
    ("title", "标题"),
    ("product_name", "产品"),
    ("severity", "严重程度"),
    ("status", "状态"),
    ("resolver", "禅道解决者"),
    ("affect_version", "影响版本"),
    ("module_name", "问题模块"),
    ("root_cause", "根因分析"),
    ("impact_scope", "影响范围"),
    ("plan_version", "计划解决版本"),
    ("issue_type_name", "问题类型"),
    ("escape_analysis", "逃逸分析"),
    ("staff_name", "解决人员"),
    ("stage", "所属阶段"),
    ("created_date", "创建时间"),
    ("remark", "备注"),
]


def export_bugs_to_excel(bugs, columns=None):
    selected_columns = BUG_EXPORT_COLUMNS
    if columns:
        selected = set(columns)
        selected_columns = [(key, title) for key, title in BUG_EXPORT_COLUMNS if key in selected]

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "网上问题"
    sheet.freeze_panes = "A2"

    for column_index, (_, title) in enumerate(selected_columns, 1):
        cell = sheet.cell(row=1, column=column_index, value=title)
        cell.font = Font(bold=True)

    for row_index, bug in enumerate(bugs, 2):
        for column_index, (key, _) in enumerate(selected_columns, 1):
            sheet.cell(row=row_index, column=column_index, value=bug.get(key, ""))

    for column_cells in sheet.columns:
        header = column_cells[0].value or ""
        width = max(12, min(max(len(str(header)), 16) + 2, 32))
        sheet.column_dimensions[column_cells[0].column_letter].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
