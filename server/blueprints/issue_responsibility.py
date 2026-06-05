from datetime import datetime
from io import BytesIO
from urllib.parse import quote

import openpyxl
from flask import Blueprint, Response, request
from openpyxl.styles import Font

from blueprints.auth import get_current_user
from blueprints.issue_admin import notify_with_dedup
from models import IssueAssessment, IssueBug, IssueProduct, IssueResponsibility, IssueStaff, IssueVersion
from utils.db import SessionLocal
from utils.permission import require_permission
from utils.response import error, paginate, success

issue_responsibility_bp = Blueprint(
    "issue_responsibility", __name__, url_prefix="/api/issue/responsibilities"
)

SEVERITY_SCORE = {1: 0.1, 2: 0.05, 3: 0.02, 4: 0}


def _responsibility_to_dict(db, record):
    staff = db.get(IssueStaff, record.staff_id)
    return {
        "id": record.id,
        "staff_id": record.staff_id,
        "staff_name": staff.name if staff else "",
        "role": record.role,
        "ratio": record.ratio,
        "description": record.description,
        "year": record.year,
    }


def _assessment_versions_query(db):
    version_ids = [item.version_id for item in db.query(IssueAssessment).all()]
    if not version_ids:
        return []
    return db.query(IssueVersion.version).filter(IssueVersion.id.in_(version_ids))


def _bug_to_resp_dict(db, bug):
    product = db.get(IssueProduct, bug.product_id) if bug.product_id else None
    records = (
        db.query(IssueResponsibility)
        .filter(IssueResponsibility.bug_id == bug.id)
        .order_by(IssueResponsibility.id)
        .all()
    )
    item = bug.to_dict()
    item["product_name"] = product.name if product else ""
    item["responsibilities"] = [_responsibility_to_dict(db, record) for record in records]
    item["has_responsibility"] = bool(records)
    return item


def _apply_resp_filters(db, query, data):
    if data.get("product_id"):
        query = query.filter(IssueBug.product_id == data["product_id"])
    assigned = data.get("assigned")
    if assigned is not None:
        assigned_ids = db.query(IssueResponsibility.bug_id)
        if assigned:
            query = query.filter(IssueBug.id.in_(assigned_ids))
        else:
            query = query.filter(~IssueBug.id.in_(assigned_ids))
    return query


@issue_responsibility_bp.route("/list", methods=["POST"])
@issue_responsibility_bp.route("/query", methods=["POST"])
@require_permission("issue/bug_view")
def resp_list():
    data = request.get_json(force=True, silent=True) or {}
    page = data.get("page", 1)
    page_size = data.get("pageSize", 20)

    db = SessionLocal()
    try:
        assessment_versions = _assessment_versions_query(db)
        if assessment_versions == []:
            return success(data=paginate([], page, page_size, 0))
        query = db.query(IssueBug).filter(IssueBug.affect_version.in_(assessment_versions))
        query = _apply_resp_filters(db, query, data)
        total = query.count()
        bugs = query.order_by(IssueBug.created_date.desc()).offset((page - 1) * page_size).limit(page_size).all()
        return success(data=paginate([_bug_to_resp_dict(db, bug) for bug in bugs], page, page_size, total))
    finally:
        db.close()


@issue_responsibility_bp.route("/save", methods=["POST"])
@require_permission("issue/resp_assign")
def resp_save():
    data = request.get_json(force=True, silent=True) or {}
    bug_id = data.get("bug_id")
    records = data.get("records", [])
    year = data.get("year") or datetime.now().year
    if not bug_id:
        return error("缺少 Bug ID")

    total_ratio = sum(float(record.get("ratio") or 0) for record in records)
    if records and abs(total_ratio - 1.0) > 0.001:
        return error("责任占比合计不为 100%")

    db = SessionLocal()
    try:
        if not db.get(IssueBug, bug_id):
            return error("问题不存在")
        staff_ids = [record.get("staff_id") for record in records if record.get("staff_id")]
        if staff_ids:
            active_count = (
                db.query(IssueStaff)
                .filter(IssueStaff.id.in_(staff_ids), IssueStaff.status == "启用")
                .count()
            )
            if active_count != len(set(staff_ids)):
                return error("责任人员不存在或已禁用")

        db.query(IssueResponsibility).filter(IssueResponsibility.bug_id == bug_id).delete()
        for record in records:
            db.add(
                IssueResponsibility(
                    bug_id=bug_id,
                    staff_id=record["staff_id"],
                    role=record.get("role", ""),
                    ratio=float(record.get("ratio") or 0),
                    description=record.get("description", ""),
                    year=year,
                )
            )
        db.commit()
        bug = db.get(IssueBug, bug_id)
        for record in records:
            staff = db.get(IssueStaff, record.get("staff_id"))
            if staff and staff.email and bug:
                notify_with_dedup(
                    "responsibility",
                    f"{bug.bug_id}:{staff.id}",
                    staff.email,
                    f"责任分配通知：{bug.bug_id}",
                    bug.title,
                )
        return success(msg="责任分配保存成功")
    finally:
        db.close()


@issue_responsibility_bp.route("/score/query", methods=["POST"])
@require_permission("issue/bug_view")
def score_query():
    data = request.get_json(force=True, silent=True) or {}
    year = data.get("year") or datetime.now().year

    db = SessionLocal()
    try:
        scores = {}
        records = db.query(IssueResponsibility).filter(IssueResponsibility.year == year).all()
        for record in records:
            bug = db.get(IssueBug, record.bug_id)
            if not bug:
                continue
            score = SEVERITY_SCORE.get(bug.severity, 0) * record.ratio
            scores.setdefault(record.staff_id, {"staff_id": record.staff_id, "raw_score": 0, "bug_count": 0})
            scores[record.staff_id]["raw_score"] += score
            scores[record.staff_id]["bug_count"] += 1

        result = []
        for staff_id, item in scores.items():
            staff = db.get(IssueStaff, staff_id)
            raw_score = round(item["raw_score"], 4)
            result.append(
                {
                    "staff_id": staff_id,
                    "staff_name": staff.name if staff else "",
                    "department": staff.department if staff else "",
                    "total_score": min(raw_score, 0.3),
                    "raw_score": raw_score,
                    "bug_count": item["bug_count"],
                }
            )
        result.sort(key=lambda row: row["total_score"], reverse=True)
        return success(data=result)
    finally:
        db.close()


@issue_responsibility_bp.route("/score/me", methods=["POST"])
@require_permission("issue/bug_view")
def score_me():
    data = request.get_json(force=True, silent=True) or {}
    year = data.get("year") or datetime.now().year
    session = get_current_user() or {}
    user_id = session.get("id") or session.get("userid")

    db = SessionLocal()
    try:
        staff = db.query(IssueStaff).filter(IssueStaff.user_id == user_id).first() if user_id else None
        if not staff:
            return success(data={"total_score": 0, "records": []})
        records = (
            db.query(IssueResponsibility)
            .filter(IssueResponsibility.staff_id == staff.id, IssueResponsibility.year == year)
            .all()
        )
        total_score = 0
        items = []
        for record in records:
            bug = db.get(IssueBug, record.bug_id)
            coefficient = SEVERITY_SCORE.get(bug.severity, 0) if bug else 0
            score = coefficient * record.ratio
            total_score += score
            items.append(
                {
                    "bug_id": bug.bug_id if bug else "",
                    "title": bug.title if bug else "",
                    "severity": bug.severity if bug else 4,
                    "ratio": record.ratio,
                    "coefficient": coefficient,
                    "score": round(score, 4),
                }
            )
        return success(data={"total_score": min(round(total_score, 4), 0.3), "records": items})
    finally:
        db.close()


@issue_responsibility_bp.route("/export", methods=["POST"])
@require_permission("issue/bug_export")
def resp_export():
    data = request.get_json(force=True, silent=True) or {}
    year = data.get("year") or datetime.now().year

    db = SessionLocal()
    try:
        rows = []
        records = (
            db.query(IssueResponsibility)
            .filter(IssueResponsibility.year == year)
            .order_by(IssueResponsibility.bug_id, IssueResponsibility.id)
            .all()
        )
        for record in records:
            bug = db.get(IssueBug, record.bug_id)
            staff = db.get(IssueStaff, record.staff_id)
            product = db.get(IssueProduct, bug.product_id) if bug and bug.product_id else None
            coefficient = SEVERITY_SCORE.get(bug.severity, 0) if bug else 0
            rows.append(
                [
                    bug.bug_id if bug else "",
                    bug.title if bug else "",
                    product.name if product else "",
                    bug.affect_version if bug else "",
                    bug.severity if bug else "",
                    staff.name if staff else "",
                    record.role,
                    record.ratio,
                    round(coefficient * record.ratio, 4),
                    record.description,
                ]
            )
    finally:
        db.close()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "责任分配"
    headers = ["Bug 编号", "标题", "产品", "影响版本", "严重程度", "责任人", "角色", "占比", "扣分", "说明"]
    for index, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.font = Font(bold=True)
    for row_index, row in enumerate(rows, 2):
        for column_index, value in enumerate(row, 1):
            sheet.cell(row=row_index, column=column_index, value=value)

    buffer = BytesIO()
    workbook.save(buffer)
    filename = "责任分配报表.xlsx"
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )
