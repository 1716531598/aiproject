from datetime import datetime, time
from io import BytesIO
from urllib.parse import quote

import openpyxl
from flask import Blueprint, Response, request
from openpyxl.styles import Font

from blueprints.auth import get_current_user
from models import (
    IssueBug,
    IssueOverdueAnalysis,
    IssuePocProject,
    IssueProduct,
    IssueResponsibility,
    IssueStaff,
    IssueTodo,
    IssueType,
)
from utils.db import SessionLocal
from utils.permission import require_permission
from utils.response import error, success

issue_statistic_bp = Blueprint("issue_statistic", __name__, url_prefix="/api/issue/statistics")

SEVERITY_SCORE = {1: 0.1, 2: 0.05, 3: 0.02, 4: 0}
RESOLVED_STATUSES = {"已解决", "已关闭"}
ACTIVE_STATUS = "激活"
BLANK_VERSION_LABEL = "未填写影响版本"


def _parse_date(value, end_of_day=False):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    try:
        parsed = datetime.fromisoformat(str(value))
    except ValueError:
        return None
    if end_of_day and parsed.time() == time.min:
        return datetime.combine(parsed.date(), time.max)
    return parsed


def _filtered_bug_query(db, data):
    query = db.query(IssueBug)
    start_date = _parse_date(data.get("start_date"))
    end_date = _parse_date(data.get("end_date"), end_of_day=True)
    if start_date:
        query = query.filter(IssueBug.created_date >= start_date)
    if end_date:
        query = query.filter(IssueBug.created_date <= end_date)
    if data.get("product_id"):
        query = query.filter(IssueBug.product_id == data["product_id"])
    return query


def _bug_stat_summary(bugs):
    total = len(bugs)
    resolved = len([bug for bug in bugs if bug.status in RESOLVED_STATUSES])
    active = len([bug for bug in bugs if bug.status == ACTIVE_STATUS])
    return {
        "total": total,
        "resolved": resolved,
        "active": active,
        "resolve_rate": round(resolved / total, 4) if total else 0,
    }


def _month_label(value):
    if not value:
        return "未填写日期"
    return value.strftime("%Y-%m")


def _bug_age_days(bug):
    if not bug.created_date:
        return 0
    now = datetime.now(bug.created_date.tzinfo) if bug.created_date.tzinfo else datetime.now()
    return max((now - bug.created_date).days, 0)


def _overdue_rows(db):
    product_map = {product.id: product.name for product in db.query(IssueProduct).all()}
    analysis_map = {
        (item.product_id, item.version): item
        for item in db.query(IssueOverdueAnalysis).all()
    }
    groups = {}
    active_bugs = db.query(IssueBug).filter(IssueBug.status == ACTIVE_STATUS).all()
    for bug in active_bugs:
        version = (bug.affect_version or "").strip()
        if not bug.product_id or not version:
            continue
        days = _bug_age_days(bug)
        if days < 30:
            continue
        key = (bug.product_id, version)
        groups.setdefault(
            key,
            {
                "product_id": bug.product_id,
                "product_name": product_map.get(bug.product_id, ""),
                "version": version,
                "overdue_total": 0,
                "overdue_180": 0,
                "overdue_360": 0,
                "bugs": [],
            },
        )
        groups[key]["overdue_total"] += 1
        if days >= 180:
            groups[key]["overdue_180"] += 1
        if days >= 360:
            groups[key]["overdue_360"] += 1
        groups[key]["bugs"].append(
            {
                "id": bug.id,
                "bug_id": bug.bug_id,
                "title": bug.title,
                "severity": bug.severity,
                "status": bug.status,
                "created_date": bug.created_date.isoformat() if bug.created_date else None,
                "overdue_days": days,
            }
        )

    result = []
    for key, row in groups.items():
        analysis = analysis_map.get(key)
        row["analysis"] = analysis.analysis if analysis else ""
        row["improvement"] = analysis.improvement if analysis else ""
        row["analysis_id"] = analysis.id if analysis else None
        row["bugs"].sort(key=lambda item: item["overdue_days"], reverse=True)
        result.append(row)
    result.sort(key=lambda item: item["overdue_total"], reverse=True)
    return result


@issue_statistic_bp.route("/overview", methods=["POST"])
@require_permission("issue/stat_view")
def overview():
    data = request.get_json(force=True, silent=True) or {}
    db = SessionLocal()
    try:
        bug_summary = _bug_stat_summary(_filtered_bug_query(db, data).all())
        poc_total = db.query(IssuePocProject).count()
        poc_closed = (
            db.query(IssuePocProject)
            .filter(IssuePocProject.current_status.like("%完成%"))
            .count()
        )
        return success(
            data={
                "bug_total": bug_summary["total"],
                "bug_resolved": bug_summary["resolved"],
                "bug_active": bug_summary["active"],
                "resolve_rate": bug_summary["resolve_rate"],
                "poc_total": poc_total,
                "poc_closed": poc_closed,
                "poc_active": poc_total - poc_closed,
            }
        )
    finally:
        db.close()


@issue_statistic_bp.route("/by-product", methods=["POST"])
@require_permission("issue/stat_view")
def by_product():
    data = request.get_json(force=True, silent=True) or {}
    db = SessionLocal()
    try:
        bugs = _filtered_bug_query(db, data).all()
        product_map = {product.id: product.name for product in db.query(IssueProduct).all()}
        grouped = {}
        for bug in bugs:
            key = bug.product_id or 0
            grouped.setdefault(key, []).append(bug)

        result = []
        for product_id, items in grouped.items():
            summary = _bug_stat_summary(items)
            result.append(
                {
                    "product_id": product_id or None,
                    "product_name": product_map.get(product_id, "未关联产品"),
                    **summary,
                }
            )
        result.sort(key=lambda item: item["total"], reverse=True)
        return success(data=result)
    finally:
        db.close()


@issue_statistic_bp.route("/by-version", methods=["POST"])
@require_permission("issue/stat_view")
def by_version():
    data = request.get_json(force=True, silent=True) or {}
    db = SessionLocal()
    try:
        bugs = _filtered_bug_query(db, data).all()
        product_map = {product.id: product.name for product in db.query(IssueProduct).all()}
        grouped = {}
        for bug in bugs:
            version = (bug.affect_version or "").strip() or BLANK_VERSION_LABEL
            key = (bug.product_id or 0, version)
            grouped.setdefault(key, []).append(bug)

        result = []
        for (product_id, version), items in grouped.items():
            summary = _bug_stat_summary(items)
            result.append(
                {
                    "product_id": product_id or None,
                    "product_name": product_map.get(product_id, "未关联产品"),
                    "version": version,
                    **summary,
                }
            )
        result.sort(key=lambda item: item["total"], reverse=True)
        return success(data=result)
    finally:
        db.close()


@issue_statistic_bp.route("/by-resolver", methods=["POST"])
@require_permission("issue/stat_view")
def by_resolver():
    data = request.get_json(force=True, silent=True) or {}
    db = SessionLocal()
    try:
        bugs = _filtered_bug_query(db, data).all()
        staff_map = {staff.id: staff.name for staff in db.query(IssueStaff).all()}
        grouped = {}
        for bug in bugs:
            key = bug.staff_id or 0
            grouped.setdefault(key, {"bugs": [], "fallback_name": bug.resolver or "未分配"})
            grouped[key]["bugs"].append(bug)

        result = []
        for staff_id, group in grouped.items():
            summary = _bug_stat_summary(group["bugs"])
            result.append(
                {
                    "staff_id": staff_id or None,
                    "staff_name": staff_map.get(staff_id, group["fallback_name"]),
                    **summary,
                }
            )
        result.sort(key=lambda item: item["total"], reverse=True)
        return success(data=result)
    finally:
        db.close()


@issue_statistic_bp.route("/by-type", methods=["POST"])
@require_permission("issue/stat_view")
def by_type():
    data = request.get_json(force=True, silent=True) or {}
    db = SessionLocal()
    try:
        bugs = _filtered_bug_query(db, data).all()
        type_map = {issue_type.id: issue_type.name for issue_type in db.query(IssueType).all()}
        grouped = {}
        for bug in bugs:
            key = bug.issue_type_id or 0
            grouped.setdefault(key, []).append(bug)

        total = len(bugs)
        result = []
        for issue_type_id, items in grouped.items():
            result.append(
                {
                    "issue_type_id": issue_type_id or None,
                    "type_name": type_map.get(issue_type_id, "未分类"),
                    "total": len(items),
                    "ratio": round(len(items) / total, 4) if total else 0,
                }
            )
        result.sort(key=lambda item: item["total"], reverse=True)
        return success(data=result)
    finally:
        db.close()


@issue_statistic_bp.route("/resolve-trend", methods=["POST"])
@require_permission("issue/stat_view")
def resolve_trend():
    data = request.get_json(force=True, silent=True) or {}
    db = SessionLocal()
    try:
        bugs = _filtered_bug_query(db, data).all()
        product_map = {product.id: product.name for product in db.query(IssueProduct).all()}
        grouped = {}
        for bug in bugs:
            key = (_month_label(bug.created_date), bug.product_id or 0)
            grouped.setdefault(key, []).append(bug)

        result = []
        for (month, product_id), items in grouped.items():
            summary = _bug_stat_summary(items)
            result.append(
                {
                    "month": month,
                    "product_id": product_id or None,
                    "product_name": product_map.get(product_id, "未关联产品"),
                    "total": summary["total"],
                    "resolved": summary["resolved"],
                    "resolve_rate": summary["resolve_rate"],
                }
            )
        result.sort(key=lambda item: (item["month"], item["product_name"]))
        return success(data=result)
    finally:
        db.close()


@issue_statistic_bp.route("/overdue", methods=["POST"])
@require_permission("issue/stat_view")
def overdue():
    db = SessionLocal()
    try:
        return success(data=_overdue_rows(db))
    finally:
        db.close()


@issue_statistic_bp.route("/overdue/analysis", methods=["POST"])
@require_permission("issue/stat_view")
def overdue_analysis():
    data = request.get_json(force=True, silent=True) or {}
    product_id = data.get("product_id")
    version = (data.get("version") or "").strip()
    if not product_id or not version:
        return error("缺少产品或版本")

    should_save = "analysis" in data or "improvement" in data
    db = SessionLocal()
    try:
        record = (
            db.query(IssueOverdueAnalysis)
            .filter(IssueOverdueAnalysis.product_id == product_id, IssueOverdueAnalysis.version == version)
            .first()
        )
        if should_save:
            if not record:
                record = IssueOverdueAnalysis(product_id=product_id, version=version)
                db.add(record)
            record.analysis = data.get("analysis", record.analysis) or ""
            record.improvement = data.get("improvement", record.improvement) or ""
            db.commit()
            db.refresh(record)
        return success(data=record.to_dict() if record else None, msg="保存成功" if should_save else "")
    finally:
        db.close()


@issue_statistic_bp.route("/overdue/export", methods=["POST"])
@require_permission("issue/stat_export")
def overdue_export():
    db = SessionLocal()
    try:
        rows = _overdue_rows(db)
    finally:
        db.close()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "超期问题"
    headers = ["产品", "版本", "超期问题总数", "超期180天", "超期360天", "问题分析", "改进措施"]
    for index, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.font = Font(bold=True)
    for row_index, row in enumerate(rows, 2):
        values = [
            row["product_name"],
            row["version"],
            row["overdue_total"],
            row["overdue_180"],
            row["overdue_360"],
            row["analysis"],
            row["improvement"],
        ]
        for column_index, value in enumerate(values, 1):
            sheet.cell(row=row_index, column=column_index, value=value)

    buffer = BytesIO()
    workbook.save(buffer)
    filename = "超期问题汇总.xlsx"
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@issue_statistic_bp.route("/my-summary", methods=["POST"])
@require_permission("issue/bug_view")
def my_summary():
    data = request.get_json(force=True, silent=True) or {}
    year = data.get("year") or datetime.now().year
    session = get_current_user() or {}
    user_id = session.get("id") or session.get("userid")

    db = SessionLocal()
    try:
        staff = db.query(IssueStaff).filter(IssueStaff.user_id == user_id).first() if user_id else None
        if not staff:
            return success(data={"bugs": [], "todos": [], "score": {"total": 0, "bug_count": 0}, "bug_stats": {"total": 0, "resolved": 0, "pending": 0}})

        active_bugs = (
            db.query(IssueBug)
            .filter(IssueBug.staff_id == staff.id, IssueBug.status == ACTIVE_STATUS)
            .order_by(IssueBug.created_date.desc())
            .limit(20)
            .all()
        )
        all_bugs = db.query(IssueBug).filter(IssueBug.staff_id == staff.id).all()
        bug_stats = {
            "total": len(all_bugs),
            "resolved": len([bug for bug in all_bugs if bug.status in RESOLVED_STATUSES]),
            "pending": len([bug for bug in all_bugs if bug.status == ACTIVE_STATUS]),
        }

        todos = (
            db.query(IssueTodo)
            .filter(IssueTodo.staff_id == staff.id, IssueTodo.status != "已完成")
            .order_by(IssueTodo.deadline.asc())
            .limit(20)
            .all()
        )
        records = (
            db.query(IssueResponsibility)
            .filter(IssueResponsibility.staff_id == staff.id, IssueResponsibility.year == year)
            .all()
        )
        total_score = 0
        for record in records:
            bug = db.get(IssueBug, record.bug_id)
            total_score += SEVERITY_SCORE.get(bug.severity, 0) * record.ratio if bug else 0

        return success(
            data={
                "bugs": [bug.to_dict() for bug in active_bugs],
                "bug_stats": bug_stats,
                "todos": [
                    {
                        "id": todo.id,
                        "content": todo.content,
                        "deadline": todo.deadline.isoformat() if todo.deadline else None,
                        "status": todo.status,
                        "project_id": todo.project_id,
                    }
                    for todo in todos
                ],
                "score": {"total": min(round(total_score, 4), 0.3), "bug_count": len(records)},
            }
        )
    finally:
        db.close()


@issue_statistic_bp.route("/score-ranking", methods=["POST"])
@require_permission("issue/bug_view")
def score_ranking():
    data = request.get_json(force=True, silent=True) or {}
    year = data.get("year") or datetime.now().year

    db = SessionLocal()
    try:
        staff_scores = {}
        records = db.query(IssueResponsibility).filter(IssueResponsibility.year == year).all()
        for record in records:
            bug = db.get(IssueBug, record.bug_id)
            score = SEVERITY_SCORE.get(bug.severity, 0) * record.ratio if bug else 0
            staff_scores.setdefault(record.staff_id, 0)
            staff_scores[record.staff_id] += score

        result = []
        for staff_id, score in sorted(staff_scores.items(), key=lambda item: item[1], reverse=True)[:10]:
            staff = db.get(IssueStaff, staff_id)
            result.append({"staff_id": staff_id, "staff_name": staff.name if staff else "", "total_score": min(round(score, 4), 0.3)})
        return success(data=result)
    finally:
        db.close()
